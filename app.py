"""
GA Attainment Automation - Streamlit App
==========================================
Upload the GA Attainment master workbook (.xlsx), enter CLO scores through
simple dropdowns instead of hunting for the right column by hand, then
review GA (Graduate Attribute) attainment analysis and charts, and download
the updated workbook.

Run locally:
    pip install streamlit pandas openpyxl plotly
    streamlit run app.py

Deploy on Streamlit Community Cloud:
    1. Push this file (and the .ipynb / README.md if you like) to a GitHub repo.
    2. On https://share.streamlit.io, click "New app", pick the repo/branch,
       and set the main file path to app.py.
    3. Streamlit Cloud installs packages from a requirements.txt in the repo
       root - add one with: streamlit, pandas, openpyxl, plotly
"""

import io
import re
from datetime import datetime

import openpyxl
import pandas as pd
import plotly.express as px
import streamlit as st

st.set_page_config(page_title="GA Attainment Automation", layout="wide")

# ---------------------------------------------------------------------------
# Core parsing / mapping logic
# ---------------------------------------------------------------------------
# These sheets store one row per student and hundreds of columns (one block
# per course, split into CLOs). The header block always follows the same
# relative layout once you find the "Student Name" cell:
#   header_row      -> "Student Name"
#   header_row + 1   -> Course name (merged across its columns)
#   header_row + 2   -> CLO label (CLO 1, CLO 2, ...)
#   header_row + 3   -> Component (Theory / Lab), optional
#   header_row + 4   -> GA label (GA1, GA2, ...) the CLO feeds into
#   header_row + 5   -> first row of student data
# Roll number sits one column to the left of the Student Name column.


def norm(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


def norm_clo(s):
    return norm(s).replace(" ", "")


def ffill(vals):
    out, last = [], None
    for v in vals:
        if v is not None:
            last = v
        out.append(last)
    return out


def find_header_anchor(ws):
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if cell.value and norm(cell.value) == "student name":
                return cell.row, cell.column
    return None


def build_maps(ws):
    """Build column map (course, component, clo) -> col index, row map roll -> row,
    and column metadata (course/component/clo/GA) for every mappable column."""
    anchor = find_header_anchor(ws)
    if anchor is None:
        return None
    name_row, name_col = anchor
    course_row, clo_row, comp_row, ga_row = (
        name_row + 1,
        name_row + 2,
        name_row + 3,
        name_row + 4,
    )
    data_start_row = name_row + 5
    roll_col = name_col - 1
    max_col, max_row = ws.max_column, ws.max_row

    course_vals = [ws.cell(row=course_row, column=c).value for c in range(1, max_col + 1)]
    clo_vals = [ws.cell(row=clo_row, column=c).value for c in range(1, max_col + 1)]
    comp_vals = [ws.cell(row=comp_row, column=c).value for c in range(1, max_col + 1)]
    ga_vals = [ws.cell(row=ga_row, column=c).value for c in range(1, max_col + 1)]

    course_ff = ffill(course_vals)
    comp_ff, last_comp, last_course = [], None, None
    for i in range(len(course_ff)):
        if course_ff[i] != last_course:
            last_comp = None
            last_course = course_ff[i]
        if comp_vals[i] is not None:
            last_comp = comp_vals[i]
        comp_ff.append(last_comp)

    col_map, col_meta = {}, {}
    for i in range(len(clo_vals)):
        col_idx = i + 1
        if clo_vals[i] is None or course_ff[i] is None:
            continue
        key = (norm(course_ff[i]), norm(comp_ff[i]) or "", norm_clo(clo_vals[i]))
        col_map.setdefault(key, col_idx)
        col_meta[col_idx] = dict(
            course=course_ff[i], component=comp_ff[i] or "", clo=clo_vals[i], ga=ga_vals[i]
        )

    row_map = {}
    for r in range(data_start_row, max_row + 1):
        roll = ws.cell(row=r, column=roll_col).value
        name = ws.cell(row=r, column=name_col).value
        if roll:
            row_map[norm(roll)] = dict(row=r, roll=roll, name=name)

    return dict(
        col_map=col_map,
        col_meta=col_meta,
        row_map=row_map,
        name_row=name_row,
        name_col=name_col,
        roll_col=roll_col,
        data_start_row=data_start_row,
    )


def list_reference_rows(maps):
    """Ordered, de-duplicated (course, component, clo) list for a sheet's dropdowns."""
    seen, out = set(), []
    for col_idx in sorted(maps["col_meta"].keys()):
        meta = maps["col_meta"][col_idx]
        key = (norm(meta["course"]), norm(meta["component"]), norm_clo(meta["clo"]))
        if key in seen:
            continue
        seen.add(key)
        out.append(meta)
    return out


def build_tidy_dataframe(wb, batch_sheets, maps_cache):
    records = []
    for sheet_name in batch_sheets:
        maps = maps_cache.get(sheet_name)
        if maps is None:
            continue
        ws = wb[sheet_name]
        for col_idx, meta in maps["col_meta"].items():
            for rinfo in maps["row_map"].values():
                val = ws.cell(row=rinfo["row"], column=col_idx).value
                if isinstance(val, (int, float)):
                    records.append(
                        dict(
                            batch=sheet_name,
                            roll=rinfo["roll"],
                            name=rinfo["name"],
                            course=meta["course"],
                            component=meta["component"],
                            clo=meta["clo"],
                            ga=meta["ga"],
                            score=val,
                        )
                    )
    return pd.DataFrame(records)


# ---------------------------------------------------------------------------
# Session state setup
# ---------------------------------------------------------------------------
if "wb" not in st.session_state:
    st.session_state.wb = None
    st.session_state.maps_cache = {}
    st.session_state.batch_sheets = []
    st.session_state.pending_entries = []
    st.session_state.applied_log = []

st.title("GA Attainment Automation")
st.caption("Upload the master workbook, enter scores through dropdowns, review GA analysis, download the result.")

# ---------------------------------------------------------------------------
# Upload
# ---------------------------------------------------------------------------
uploaded = st.file_uploader("Upload the GA Attainment Excel workbook (.xlsx)", type=["xlsx"])

if uploaded is not None and st.session_state.wb is None:
    with st.spinner("Reading workbook and mapping every course/CLO column..."):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(uploaded.getvalue()), data_only=False)
        except Exception as e:
            st.error(
                f"Could not open '{uploaded.name}' as an Excel workbook ({e}). "
                "Make sure you uploaded a valid, un-corrupted .xlsx file."
            )
            st.stop()

        batch_sheets = [s for s in wb.sheetnames if s.lower() != "sheet1"]
        if not batch_sheets:
            st.error(
                f"No batch sheets found in this file (only 'Sheet1' or nothing "
                f"at all). Sheets found: {wb.sheetnames}"
            )
            st.stop()

        maps_cache = {}
        unparsed_sheets = []
        for s in batch_sheets:
            m = build_maps(wb[s])
            if m is not None:
                maps_cache[s] = m
            else:
                unparsed_sheets.append(s)

        st.session_state.wb = wb
        st.session_state.batch_sheets = [s for s in batch_sheets if s in maps_cache]
        st.session_state.maps_cache = maps_cache
        st.session_state.source_filename = uploaded.name
        st.session_state.unparsed_sheets = unparsed_sheets

    st.success(f"Loaded {len(st.session_state.batch_sheets)} batch sheet(s).")
    if st.session_state.get("unparsed_sheets"):
        st.warning(
            "Could not find a 'Student Name' header in these sheet(s), so they "
            f"were skipped: {st.session_state.unparsed_sheets}"
        )

if st.button("Start over / upload a different file"):
    st.session_state.wb = None
    st.session_state.maps_cache = {}
    st.session_state.batch_sheets = []
    st.session_state.pending_entries = []
    st.session_state.applied_log = []
    st.rerun()

if st.session_state.wb is None:
    st.info("Upload a workbook to begin.")
    st.stop()

if not st.session_state.batch_sheets:
    st.error(
        "None of the sheets in this workbook could be parsed (no 'Student Name' "
        "header found anywhere). Click 'Start over' above and check the file."
    )
    st.stop()

wb = st.session_state.wb
maps_cache = st.session_state.maps_cache
batch_sheets = st.session_state.batch_sheets

tab_entry, tab_analysis, tab_download = st.tabs(
    ["Data Entry", "Analysis & Visualization", "Download"]
)

# ---------------------------------------------------------------------------
# TAB 1: Data entry
# ---------------------------------------------------------------------------
with tab_entry:
    st.subheader("Add a score")
    col1, col2 = st.columns(2)

    with col1:
        batch = st.selectbox("Batch", batch_sheets, key="de_batch")
        maps = maps_cache[batch]
        ref_rows = list_reference_rows(maps)

        courses = sorted({r["course"] for r in ref_rows})
        course = st.selectbox("Course", courses, key="de_course")

        components = sorted({r["component"] for r in ref_rows if r["course"] == course})
        component = st.selectbox("Component", components if components else [""], key="de_comp")

        clos = [
            r["clo"]
            for r in ref_rows
            if r["course"] == course and r["component"] == component
        ]
        clo = st.selectbox("CLO", clos, key="de_clo")

    with col2:
        roll_options = sorted(
            [f'{info["roll"]} - {info["name"]}' for info in maps["row_map"].values()]
        )
        roll_choice = st.selectbox("Student (Roll No - Name)", roll_options, key="de_roll")
        roll = roll_choice.split(" - ")[0] if roll_choice else ""
        score = st.number_input("Score", min_value=0.0, max_value=100.0, step=0.01, key="de_score")

    if st.button("Add to queue", type="primary"):
        st.session_state.pending_entries.append(
            dict(batch=batch, course=course, component=component, clo=clo, roll=roll, score=score)
        )
        st.success("Added to the queue below.")

    st.divider()
    st.subheader("Pending entries")
    if st.session_state.pending_entries:
        pending_df = pd.DataFrame(st.session_state.pending_entries)
        st.dataframe(pending_df, use_container_width=True)

        c1, c2 = st.columns(2)
        with c1:
            if st.button("Apply all pending entries to the workbook"):
                applied, skipped = 0, []
                for i, entry in enumerate(st.session_state.pending_entries):
                    maps = maps_cache[entry["batch"]]
                    ws = wb[entry["batch"]]
                    key = (norm(entry["course"]), norm(entry["component"]), norm_clo(entry["clo"]))
                    col_idx = maps["col_map"].get(key)
                    rinfo = maps["row_map"].get(norm(entry["roll"]))
                    if col_idx is None or rinfo is None:
                        skipped.append(entry)
                        continue
                    ws.cell(row=rinfo["row"], column=col_idx).value = entry["score"]
                    applied += 1
                    st.session_state.applied_log.append(
                        dict(**entry, applied_at=datetime.now().isoformat(timespec="seconds"))
                    )
                st.session_state.pending_entries = skipped
                st.success(f"Applied {applied} score(s) to the workbook.")
                if skipped:
                    st.warning(f"{len(skipped)} entrie(s) could not be matched and were kept in the queue.")
                st.rerun()
        with c2:
            if st.button("Clear queue"):
                st.session_state.pending_entries = []
                st.rerun()
    else:
        st.caption("No entries queued yet. Add one above.")

    if st.session_state.applied_log:
        with st.expander(f"Applied entries log ({len(st.session_state.applied_log)})"):
            st.dataframe(pd.DataFrame(st.session_state.applied_log), use_container_width=True)

# ---------------------------------------------------------------------------
# TAB 2: Analysis & visualization
# ---------------------------------------------------------------------------
with tab_analysis:
    st.subheader("GA Attainment Analysis")
    with st.spinner("Building analysis dataset..."):
        df = build_tidy_dataframe(wb, batch_sheets, maps_cache)

    if df.empty:
        st.info("No numeric scores found yet.")
    else:
        df["ga_norm"] = df["ga"].astype(str).str.replace(" ", "").str.upper()
        df = df[df["ga_norm"].str.match(r"^GA\d+$")]

        analysis_batches = st.multiselect(
            "Batches to include", batch_sheets, default=batch_sheets
        )
        view = df[df["batch"].isin(analysis_batches)] if analysis_batches else df

        attainment = (
            view.groupby(["batch", "roll", "name", "ga_norm"])["score"].mean().reset_index()
        )

        st.markdown("**Average GA attainment per batch**")
        ga_batch_avg = (
            attainment.groupby(["batch", "ga_norm"])["score"].mean().reset_index()
        )
        ga_order = sorted(ga_batch_avg["ga_norm"].unique(), key=lambda x: int(x[2:]))
        fig1 = px.bar(
            ga_batch_avg,
            x="ga_norm",
            y="score",
            color="batch",
            barmode="group",
            category_orders={"ga_norm": ga_order},
            labels={"ga_norm": "Graduate Attribute", "score": "Average attainment (%)"},
        )
        st.plotly_chart(fig1, use_container_width=True)

        st.markdown("**GA attainment heatmap (batch x GA)**")
        heat = ga_batch_avg.pivot(index="batch", columns="ga_norm", values="score")
        heat = heat[[c for c in ga_order if c in heat.columns]]
        fig2 = px.imshow(
            heat,
            color_continuous_scale="RdYlGn",
            aspect="auto",
            labels=dict(color="Attainment %"),
        )
        st.plotly_chart(fig2, use_container_width=True)

        st.markdown("**Score distribution by course (selected batch)**")
        dist_batch = st.selectbox("Batch for course-level view", analysis_batches or batch_sheets)
        course_df = view[view["batch"] == dist_batch]
        course_avg = (
            course_df.groupby("course")["score"].mean().reset_index().sort_values("score")
        )
        fig3 = px.bar(
            course_avg,
            x="score",
            y="course",
            orientation="h",
            labels={"score": "Average score", "course": "Course"},
        )
        fig3.update_layout(height=max(400, 20 * len(course_avg)))
        st.plotly_chart(fig3, use_container_width=True)

        st.markdown("**Students below a chosen attainment threshold**")
        threshold = st.slider("Threshold (%)", 0, 100, 50)
        low = attainment[attainment["score"] < threshold].sort_values("score")
        st.dataframe(low, use_container_width=True)

        st.markdown("**Raw tidy dataset (for your own analysis)**")
        st.dataframe(view, use_container_width=True)
        st.download_button(
            "Download this data as CSV",
            view.to_csv(index=False).encode("utf-8"),
            file_name="ga_scores_tidy.csv",
            mime="text/csv",
        )

# ---------------------------------------------------------------------------
# TAB 3: Download updated workbook
# ---------------------------------------------------------------------------
with tab_download:
    st.subheader("Download the updated workbook")
    st.write(
        "This saves the workbook exactly as it is in memory right now, including "
        "every entry you've applied in the Data Entry tab."
    )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    out_name = st.session_state.get("source_filename", "GA_Attainment_updated.xlsx")
    st.download_button(
        "Download updated .xlsx",
        buf,
        file_name=out_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.caption(
        "Note: formulas already in the sheet (e.g. GA average formulas) are preserved as "
        "formulas but won't show a computed value until the file is opened and recalculated "
        "in Excel or LibreOffice."
    )
